from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, Response
import sqlite3
from werkzeug.security import check_password_hash, generate_password_hash
from database import (
    get_db, get_user_by_student_id, get_user_identities, get_user_materials,
    get_transfer_status, initialize_user_materials, update_transfer_status, save_signature, get_abandoned_users,
    update_material_status, get_material_requirements, update_material_requirements,
    get_announcement, update_announcement, get_contact_info, update_contact_info,
    get_help_content, update_help_content, get_all_users, add_user, update_user,
    delete_user, get_dashboard_metrics, add_or_update_user_material,
    get_user_signature, get_user_cultivators, get_all_cultivators, add_cultivator,
    update_cultivator, delete_cultivator, get_user_cultivator_ids, update_user_cultivators,
    clean_unused_materials, delete_material, get_all_identities, update_user_identities,
    reset_user_data, get_user_identity_ids, log_operation, get_operation_logs,
    search_operation_logs, ensure_fee_records_table,
    get_fee_records, get_user_fee_records, add_fee_record, update_fee_record,
    delete_fee_record, batch_generate_fee_records, get_fee_statistics
)
import json
import logging
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# 配置日志记录器
logging.basicConfig(
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# SECRET_KEY: 优先使用环境变量，否则从文件读取/生成并持久化
_secret_key_env = os.environ.get('SECRET_KEY')
if _secret_key_env:
    app.secret_key = _secret_key_env
else:
    _secret_key_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.secret_key')
    if os.path.exists(_secret_key_file):
        with open(_secret_key_file, 'r') as f:
            app.secret_key = f.read().strip()
    else:
        app.secret_key = os.urandom(32).hex()
        with open(_secret_key_file, 'w') as f:
            f.write(app.secret_key)
        os.chmod(_secret_key_file, 0o600)

# 确保数据库表存在
from database import ensure_operation_logs_table
ensure_operation_logs_table()
ensure_fee_records_table()

# Define add_role filter
def add_role_filter(value, cultivators):
    for cultivator in cultivators:
        if cultivator['name'] == value:
            return f"{value}（{cultivator['role']}）"
    return value

app.jinja_env.filters['add_role'] = add_role_filter

def login_required(f):
    def wrap(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrap.__name__ = f.__name__
    return wrap

def admin_required(f):
    def wrap(*args, **kwargs):
        if 'user_id' not in session or session.get('is_admin') != 1:
            flash('无权限访问', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrap.__name__ = f.__name__
    return wrap

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('transfer'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        student_id = request.form['student_id']
        password = request.form['password']
        user = get_user_by_student_id(student_id)
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['student_id'] = user['student_id']
            session['is_admin'] = user['is_admin']
            logger.debug(f"User logged in: student_id={student_id}, user_id={user['id']}, is_admin={user['is_admin']}")
            if user['is_admin']:
                log_operation(user['id'], student_id, '管理员登录')
                return redirect(url_for('admin'))
            return redirect(url_for('transfer_popup'))
        flash('学号或密码错误', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('已退出登录', 'success')
    return redirect(url_for('login'))

@app.route('/help')
@login_required
def help():
    help_content = get_help_content()
    return render_template('help.html', help_content=help_content)

@app.route('/transfer_popup')
@login_required
def transfer_popup():
    user_id = session['user_id']
    logger.debug(f"Rendering transfer_popup.html for user_id={user_id}")
    identities = get_user_identities(user_id)
    has_activist = any(identity['identity_name'] == '入党积极分子' for identity in identities)
    if not has_activist:
        return redirect(url_for('transfer'))
    return render_template('transfer_popup.html')

@app.route('/transfer', methods=['GET', 'POST'])
@login_required
def transfer():
    user_id = session['user_id']
    logger.debug(f"Transfer route accessed for user_id={user_id}")
    
    initialize_user_materials(user_id)
    
    # 直接通过 user_id 查询用户信息，避免冗余数据库连接
    conn = get_db()
    try:
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    finally:
        conn.close()
    if not user:
        logger.error(f"No user found for user_id={user_id}")
        flash('用户不存在', 'error')
        return redirect(url_for('logout'))
    identities = get_user_identities(user_id)
    materials = get_user_materials(user_id)
    logger.debug(f"Materials for user_id={user_id}: {[(m['id'], m['name'], m['status'], m['details']) for m in materials]}")
    transfer_status = get_transfer_status(user_id)
    signature = get_user_signature(user_id)
    cultivators = get_user_cultivators(user_id)
    
    material_requirements = {
        identity['id']: get_material_requirements(identity['id'])
        for identity in identities
    }
    logger.debug(f"Material requirements: {material_requirements}")
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'abandon':
            has_activist = any(identity['identity_name'] == '入党积极分子' for identity in identities)
            if not has_activist:
                flash('您无权放弃转接', 'error')
                return redirect(url_for('transfer'))
            return redirect(url_for('signature'))
    
    required_material_ids = set()
    for identity in identities:
        for material in material_requirements[identity['id']]:
            required_material_ids.add(material['id'])
    total_materials = len(required_material_ids)
    complete_materials = sum(1 for m in materials if m['status'] == '齐全' and m['id'] in required_material_ids)
    completeness = (complete_materials / total_materials * 100) if total_materials > 0 else 0
    logger.debug(f"Completeness for user_id={user_id}: {complete_materials}/{total_materials} = {completeness:.1f}%")
    
    announcement = get_announcement()
    
    return render_template('transfer.html', 
                         identities=identities, 
                         materials=materials, 
                         transfer_status=transfer_status, 
                         completeness=completeness,
                         announcement=announcement,
                         current_user=user,
                         material_requirements=material_requirements,
                         signature=signature,
                         cultivators=cultivators)

@app.route('/signature', methods=['GET', 'POST'])
@login_required
def signature():
    user_id = session['user_id']
    identities = get_user_identities(user_id)
    has_activist = any(identity['identity_name'] == '入党积极分子' for identity in identities)
    if not has_activist:
        flash('您无权放弃转接', 'error')
        return redirect(url_for('transfer'))
    
    if request.method == 'POST':
        signature_data = request.form['signature']
        update_transfer_status(user_id, '已放弃', '已放弃', abandoned=True)
        save_signature(user_id, signature_data)
        log_operation(session['user_id'], session.get('student_id', '用户'), '放弃档案转接')
        flash('已放弃档案转接', 'success')
        return redirect(url_for('transfer'))
    
    return render_template('signature.html')

@app.route('/signature/<int:user_id>')
@login_required
def get_signature(user_id):
    if not session.get('is_admin') and session['user_id'] != user_id:
        flash('无权限查看该签字', 'error')
        return jsonify({'success': False, 'error': '无权限'}), 403
    signature = get_user_signature(user_id)
    if signature:
        logger.debug(f"Signature retrieved for user_id={user_id}")
        return jsonify({'success': True, 'signature_data': signature['signature_data'], 'created_at': signature['created_at']})
    logger.debug(f"No signature found for user_id={user_id}")
    return jsonify({'success': False, 'error': '无签字记录'})

@app.route('/admin')
@admin_required
def admin():
    metrics = get_dashboard_metrics()
    abandoned_users = get_abandoned_users()
    announcement = get_announcement()
    contact_info = get_contact_info()
    help_content = get_help_content()
    return render_template('admin.html', 
                         metrics=metrics, 
                         abandoned_users=abandoned_users,
                         announcement=announcement,
                         contact_info=contact_info,
                         help_content=help_content)

@app.route('/admin/users', methods=['GET', 'POST'])
@admin_required
def admin_users():
    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'add':
                student_id = request.form['student_id']
                name = request.form['name']
                password = request.form['password']
                class_name = request.form['class_name']
                batch = request.form['batch']
                is_admin = 1 if request.form.get('is_admin') == 'on' else 0
                identity_ids = [int(i) for i in request.form.getlist('identities')]
                cultivator_ids = [int(i) for i in request.form.getlist('cultivators')]
                
                user_id = add_user(student_id, name, password, class_name, batch, is_admin)
                update_user_identities(user_id, identity_ids)
                update_user_cultivators(user_id, cultivator_ids)
                log_operation(session['user_id'], session.get('student_id', '管理员'), '添加用户', f'{name}({student_id})')
                flash('用户添加成功', 'success')
                
            elif action == 'edit':
                user_id = int(request.form['user_id'])
                student_id = request.form['student_id']
                name = request.form['name']
                class_name = request.form['class_name']
                batch = request.form['batch']
                is_admin = 1 if request.form.get('is_admin') == 'on' else 0
                password = request.form.get('password')
                identity_ids = [int(i) for i in request.form.getlist('identities')]
                cultivator_ids = [int(i) for i in request.form.getlist('cultivators')]
                
                update_user(user_id, student_id, name, class_name, batch, is_admin, password)
                update_user_identities(user_id, identity_ids)
                update_user_cultivators(user_id, cultivator_ids)
                log_operation(session['user_id'], session.get('student_id', '管理员'), '修改用户', f'{name}({student_id})')
                flash('用户更新成功', 'success')
                
            elif action == 'delete':
                user_id = int(request.form['user_id'])
                delete_user(user_id)
                log_operation(session['user_id'], session.get('student_id', '管理员'), '删除用户', f'ID:{user_id}')
                flash('用户删除成功', 'success')
                
        except sqlite3.IntegrityError as e:
            flash(f'操作失败：学号可能已存在', 'error')
            logger.error(f"Database error: {str(e)}")
        except Exception as e:
            flash(f'操作失败：{str(e)}', 'error')
            logger.error(f"Unexpected error: {str(e)}")
        
        return redirect(url_for('admin_users'))
    
    users = get_all_users()
    cultivators = get_all_cultivators()
    identities = get_all_identities()
    users_with_details = []
    
    for user in users:
        identity_ids = get_user_identity_ids(user['id'])
        identity_names = [i['identity_name'] for i in get_user_identities(user['id'])]
        cultivator_ids = get_user_cultivator_ids(user['id'])
        user_dict = dict(user)
        user_dict['identity_ids'] = identity_ids
        user_dict['identity_names'] = identity_names
        user_dict['cultivator_ids'] = cultivator_ids
        users_with_details.append(user_dict)
    
    return render_template('admin_users.html', 
                         users=users_with_details, 
                         cultivators=cultivators,
                         identities=identities)

@app.route('/apply_transfer', methods=['POST'])
@login_required
def apply_transfer():
    user_id = session['user_id']
    receiver = request.form.get('receiver')
    if not receiver:
        flash('接收地不能为空', 'error')
        return redirect(url_for('transfer'))
    try:
        update_transfer_status(user_id, '处理中', '待确认', receiver=receiver, progress=25)
        log_operation(session['user_id'], session.get('student_id', '用户'), '提交转接申请', receiver)
        flash('转接申请提交成功', 'success')
    except Exception as e:
        logger.error(f"Error applying transfer for user_id={user_id}: {str(e)}")
        flash(f'提交失败：{str(e)}', 'error')
    return redirect(url_for('transfer'))

@app.route('/admin/update_transfer_status', methods=['POST'])
@admin_required
def admin_update_transfer_status():
    try:
        user_id = int(request.form['user_id'])
        transfer_status = request.form['transfer_status']
        receive_status = request.form['receive_status']
        progress = int(request.form['progress'])
        receiver = request.form.get('receiver', '')
        transfer_date = request.form.get('transfer_date', '')
        update_transfer_status(user_id, transfer_status, receive_status, 
                             receiver=receiver, progress=progress, transfer_date=transfer_date)
        log_operation(session['user_id'], session.get('student_id', '管理员'), '更新转接状态', f'用户ID:{user_id}', f'{transfer_status}/{receive_status}')
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error updating transfer status for user_id={user_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/transfer_status')
@admin_required
def admin_transfer_status():
    try:
        conn = get_db()
        users = conn.execute('''
            SELECT u.id, u.student_id, u.name, u.class_name, u.batch,
                   GROUP_CONCAT(i.name) as identity_names,
                   ts.transfer_status, ts.receive_status, ts.receiver, ts.progress, ts.transfer_date
            FROM users u
            LEFT JOIN user_identities ui ON u.id = ui.user_id
            LEFT JOIN identities i ON ui.identity_id = i.id
            LEFT JOIN transfer_status ts ON u.id = ts.user_id
            WHERE u.is_admin = 0
            GROUP BY u.id
            ORDER BY u.id
        ''').fetchall()
        identities = conn.execute('SELECT DISTINCT name FROM identities').fetchall()
        conn.close()
        users = [dict(user) for user in users]
        for user in users:
            user['identity_names'] = user['identity_names'].split(',') if user['identity_names'] else []
        return render_template('admin_transfer_status.html', 
                             users=users, 
                             identities=[i['name'] for i in identities])
    except Exception as e:
        logger.error(f"Error fetching transfer statuses: {str(e)}")
        flash(f'加载失败：{str(e)}', 'error')
        return redirect(url_for('admin'))

@app.route('/admin/reset_users', methods=['GET', 'POST'])
@admin_required
def reset_users():
    if request.method == 'POST':
        user_ids = request.form.getlist('user_ids')
        try:
            if not user_ids:
                flash('请至少选择一个用户进行初始化', 'error')
                return redirect(url_for('reset_users'))
            for user_id in user_ids:
                reset_user_data(user_id)
            log_operation(session['user_id'], session.get('student_id', '管理员'), '初始化用户数据', f'{len(user_ids)}个用户')
            flash(f'已成功初始化 {len(user_ids)} 个用户的数据', 'success')
        except Exception as e:
            flash(f'初始化失败：{str(e)}', 'error')
            logger.error(f"Error resetting users: {str(e)}")
        return redirect(url_for('admin_users'))
    
    users = get_all_users()
    return render_template('reset_users.html', users=users)

@app.route('/admin/materials', methods=['GET', 'POST'])
@admin_required
def admin_materials():
    if request.method == 'POST':
        try:
            identity_id = int(request.form['identity_id'])
            materials = [m.strip() for m in request.form.getlist('materials[]') if m.strip()]
            update_material_requirements(identity_id, materials)
            clean_unused_materials()
            flash('材料要求更新成功', 'success')
        except ValueError:
            flash('无效的身份ID', 'error')
            logger.error(f"Invalid identity_id: {request.form['identity_id']}")
        except Exception as e:
            flash(f'更新材料要求失败: {str(e)}', 'error')
            logger.error(f"Error updating material requirements: {str(e)}")
        return redirect(url_for('admin_materials'))
    
    identities = get_all_identities()
    material_requirements = {identity['id']: get_material_requirements(identity['id']) for identity in identities}
    return render_template('admin_materials.html', 
                         identities=identities, 
                         material_requirements=material_requirements)

@app.route('/admin/delete_material', methods=['POST'])
@admin_required
def delete_material_route():
    try:
        material_id = int(request.form['material_id'])
        identity_id = int(request.form['identity_id'])
        result = delete_material(material_id, identity_id)
        if result['success']:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': result['error']}), 400
    except ValueError:
        logger.error(f"Invalid material_id: {request.form['material_id']}, identity_id: {request.form['identity_id']}")
        return jsonify({'success': False, 'error': '无效的材料ID或身份ID'}), 400
    except Exception as e:
        logger.error(f"Error deleting material: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/material_checklist')
@admin_required
def material_checklist():
    identities = get_all_identities()
    return render_template('admin_material_checklist.html', identities=identities)

@app.route('/admin/api/material_checklist', methods=['GET'])
@admin_required
def api_material_checklist():
    try:
        search = request.args.get('search', '').strip()
        batch = request.args.get('batch', '').strip()
        identity_id = request.args.get('identity_id', '')
        completeness = request.args.get('completeness', 'all')

        conn = get_db()
        with conn:
            where_clauses = ["u.is_admin = 0"]
            params = []
            
            if search:
                where_clauses.append("(u.student_id LIKE ? OR u.name LIKE ? OR u.class_name LIKE ?)")
                search_term = f"%{search}%"
                params.extend([search_term, search_term, search_term])
            
            if batch:
                where_clauses.append("u.batch LIKE ?")
                params.append(f"%{batch}%")
            
            if identity_id:
                where_clauses.append("ui.identity_id = ?")
                params.append(int(identity_id))
            
            where_clause = " AND ".join(where_clauses) if where_clauses else "1=1"
            
            query = f'''
                WITH UserMaterials AS (
                    SELECT um.user_id, m.id AS material_id, m.name AS material_name,
                           COALESCE(um.status, '') AS status,
                           COALESCE(um.details, '') AS details,
                           GROUP_CONCAT(im.identity_id) AS identity_ids,
                           (SELECT COUNT(*) FROM material_images mi WHERE mi.user_id = um.user_id AND mi.material_id = m.id) AS image_count
                    FROM materials m
                    LEFT JOIN user_materials um ON m.id = um.material_id
                    LEFT JOIN identity_materials im ON m.id = im.material_id
                    GROUP BY m.id, um.user_id
                ),
                UserCompleteness AS (
                    SELECT um.user_id,
                           CASE
                               WHEN COUNT(um.status) = 0 THEN 'incomplete'
                               WHEN SUM(CASE WHEN um.status = '齐全' THEN 1 ELSE 0 END) = COUNT(um.status) THEN 'complete'
                               WHEN SUM(CASE WHEN um.status = '待审核' THEN 1 ELSE 0 END) = COUNT(um.status) THEN 'pending'
                               ELSE 'partial'
                           END AS completeness
                    FROM UserMaterials um
                    GROUP BY um.user_id
                )
                SELECT u.id, u.student_id, u.name, u.batch, u.class_name,
                       GROUP_CONCAT(i.id) AS identity_ids,
                       GROUP_CONCAT(i.name) AS identity_names,
                       uc.completeness,
                       ts.transfer_status,
                       (SELECT json_group_array(
                           json_object(
                               'id', um.material_id,
                               'name', um.material_name,
                               'status', um.status,
                               'details', um.details,
                               'identity_ids', um.identity_ids,
                               'image_count', um.image_count
                           )
                       ) FROM UserMaterials um WHERE um.user_id = u.id) AS materials
                FROM users u
                LEFT JOIN user_identities ui ON u.id = ui.user_id
                LEFT JOIN identities i ON ui.identity_id = i.id
                LEFT JOIN UserCompleteness uc ON u.id = uc.user_id
                LEFT JOIN transfer_status ts ON u.id = ts.user_id
                WHERE {where_clause}
                {'AND uc.completeness = ?' if completeness != 'all' else ''}
                GROUP BY u.id
                ORDER BY u.id
            '''
            if completeness != 'all':
                params.append(completeness)
            
            users = conn.execute(query, params).fetchall()
            
            count_query = f'''
                SELECT COUNT(DISTINCT u.id)
                FROM users u
                LEFT JOIN user_identities ui ON u.id = ui.user_id
                WHERE {where_clause}
            '''
            count_params = params[:-1] if completeness != 'all' else params
            total_users = conn.execute(count_query, count_params).fetchone()[0]
            
            materials = conn.execute('''
                SELECT m.id, m.name, GROUP_CONCAT(im.identity_id) AS identity_ids
                FROM materials m
                LEFT JOIN identity_materials im ON m.id = im.material_id
                GROUP BY m.id
                ORDER BY m.id
            ''').fetchall()

        users_data = []
        for user in users:
            user_dict = dict(user)
            user_dict['identity_ids'] = [int(i) for i in (user['identity_ids'].split(',') if user['identity_ids'] else [])]
            user_dict['identity_names'] = user['identity_names'].split(',') if user['identity_names'] else []
            user_dict['materials'] = json.loads(user['materials'])
            for material in user_dict['materials']:
                material['identity_ids'] = [int(i) for i in (material['identity_ids'].split(',') if material['identity_ids'] else [])]
            users_data.append(user_dict)
        
        materials_data = [dict(m) for m in materials]
        for m in materials_data:
            m['identity_ids'] = [int(i) for i in (m['identity_ids'].split(',') if m['identity_ids'] else [])]
        
        return jsonify({
            'success': True,
            'users': users_data,
            'materials': materials_data,
            'identities': [dict(i) for i in get_all_identities()],
            'total_users': total_users
        })
    except Exception as e:
        logger.error(f"Error fetching material checklist: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/api/update_user_material', methods=['POST'])
@admin_required
def update_user_material():
    try:
        data = request.form
        user_id = int(data['user_id'])
        material_id = int(data['material_id'])
        status = data['status']
        details = data.get('details', '')

        conn = get_db()
        with conn:
            exists = conn.execute('''
                SELECT 1 FROM user_materials WHERE user_id = ? AND material_id = ?
            ''', (user_id, material_id)).fetchone()
            if exists:
                conn.execute('''
                    UPDATE user_materials
                    SET status = ?, details = ?
                    WHERE user_id = ? AND material_id = ?
                ''', (status, details, user_id, material_id))
            else:
                conn.execute('''
                    INSERT INTO user_materials (user_id, material_id, status, details)
                    VALUES (?, ?, ?, ?)
                ''', (user_id, material_id, status, details))
        logger.debug(f"Updated material for user_id={user_id}, material_id={material_id}: status={status}")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error updating material for user_id={user_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/api/upload_material_image', methods=['POST'])
@admin_required
def upload_material_image():
    try:
        user_id = int(request.form['user_id'])
        material_id = int(request.form['material_id'])
        if 'file' not in request.files or not request.files['file']:
            return jsonify({'success': False, 'error': '未选择文件'}), 400
        
        file = request.files['file']
        if file.mimetype not in ['image/png', 'image/jpeg', 'application/pdf']:
            return jsonify({'success': False, 'error': '仅支持 PNG、JPEG 或 PDF 格式'}), 400
        
        file_data = file.read()
        if len(file_data) > 5 * 1024 * 1024:  # Limit to 5MB
            return jsonify({'success': False, 'error': '文件大小不能超过 5MB'}), 400

        conn = get_db()
        with conn:
            conn.execute('''
                INSERT INTO material_images (user_id, material_id, file_data)
                VALUES (?, ?, ?)
            ''', (user_id, material_id, file_data))
        
        logger.debug(f"Uploaded image for user_id={user_id}, material_id={material_id}")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error uploading image for user_id={user_id}, material_id={material_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/api/material_image/<int:user_id>/<int:material_id>/<int:image_id>')
@admin_required
def get_material_image(user_id, material_id, image_id):
    try:
        conn = get_db()
        try:
            result = conn.execute('''
                SELECT file_data FROM material_images
                WHERE user_id = ? AND material_id = ? AND id = ?
            ''', (user_id, material_id, image_id)).fetchone()
        finally:
            conn.close()
        
        if not result:
            return jsonify({'success': False, 'error': '未找到图片'}), 404
        
        file_data = result['file_data']
        if file_data.startswith(b'\xff\xd8'):
            mime_type = 'image/jpeg'
        elif file_data.startswith(b'\x89PNG'):
            mime_type = 'image/png'
        elif file_data.startswith(b'%PDF'):
            mime_type = 'application/pdf'
        else:
            mime_type = 'application/octet-stream'
        return Response(file_data, mimetype=mime_type)
    except Exception as e:
        logger.error(f"Error retrieving image for user_id={user_id}, material_id={material_id}, image_id={image_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/api/material_images/<int:user_id>/<int:material_id>', methods=['GET'])
@admin_required
def get_material_images(user_id, material_id):
    try:
        conn = get_db()
        images = conn.execute('''
            SELECT id, created_at
            FROM material_images
            WHERE user_id = ? AND material_id = ?
            ORDER BY created_at
        ''', (user_id, material_id)).fetchall()
        conn.close()
        
        images_data = [{'id': img['id'], 'created_at': img['created_at']} for img in images]
        return jsonify({'success': True, 'images': images_data})
    except Exception as e:
        logger.error(f"Error retrieving images for user_id={user_id}, material_id={material_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/api/delete_material_image', methods=['POST'])
@admin_required
def delete_material_image():
    try:
        image_id = int(request.form['image_id'])
        conn = get_db()
        with conn:
            result = conn.execute('''
                DELETE FROM material_images WHERE id = ?
            ''', (image_id,)).rowcount
        if result == 0:
            return jsonify({'success': False, 'error': '图片不存在'}), 404
        logger.debug(f"Deleted image_id={image_id}")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting image_id={image_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/archive_overview')
@admin_required
def archive_overview():
    identities = get_all_identities()
    return render_template('admin_archive_overview.html', identities=identities)

@app.route('/admin/api/export_archive', methods=['GET'])
@admin_required
def export_archive():
    try:
        search = request.args.get('search', '').strip()
        batch = request.args.get('batch', '').strip()
        identity_id = request.args.get('identity_id', '')
        completeness = request.args.get('completeness', 'all')

        conn = get_db()
        with conn:
            where_clauses = ["u.is_admin = 0"]
            params = []
            
            if search:
                where_clauses.append("(u.student_id LIKE ? OR u.name LIKE ? OR u.class_name LIKE ?)")
                search_term = f"%{search}%"
                params.extend([search_term, search_term, search_term])
            
            if batch:
                where_clauses.append("u.batch LIKE ?")
                params.append(f"%{batch}%")
            
            if identity_id:
                where_clauses.append("ui.identity_id = ?")
                params.append(int(identity_id))
            
            where_clause = " AND ".join(where_clauses) if where_clauses else "1=1"
            
            query = f'''
                WITH UserMaterials AS (
                    SELECT um.user_id, m.id AS material_id, m.name AS material_name,
                           COALESCE(um.status, '') AS status,
                           COALESCE(um.details, '') AS details,
                           GROUP_CONCAT(im.identity_id) AS identity_ids
                    FROM materials m
                    LEFT JOIN user_materials um ON m.id = um.material_id
                    LEFT JOIN identity_materials im ON m.id = im.material_id
                    GROUP BY m.id, um.user_id
                ),
                UserCompleteness AS (
                    SELECT um.user_id,
                           CASE
                               WHEN COUNT(um.status) = 0 THEN 'incomplete'
                               WHEN SUM(CASE WHEN um.status = '齐全' THEN 1 ELSE 0 END) = COUNT(um.status) THEN 'complete'
                               WHEN SUM(CASE WHEN um.status = '待审核' THEN 1 ELSE 0 END) = COUNT(um.status) THEN 'pending'
                               ELSE 'partial'
                           END AS completeness
                    FROM UserMaterials um
                    GROUP BY um.user_id
                )
                SELECT u.id, u.student_id, u.name, u.batch, u.class_name,
                       GROUP_CONCAT(i.name) AS identity_names,
                       uc.completeness,
                       ts.transfer_status
                FROM users u
                LEFT JOIN user_identities ui ON u.id = ui.user_id
                LEFT JOIN identities i ON ui.identity_id = i.id
                LEFT JOIN UserCompleteness uc ON u.id = uc.user_id
                LEFT JOIN transfer_status ts ON u.id = ts.user_id
                WHERE {where_clause}
                {'AND uc.completeness = ?' if completeness != 'all' else ''}
                GROUP BY u.id
                ORDER BY u.id
            '''
            if completeness != 'all':
                params.append(completeness)
            
            users = conn.execute(query, params).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "档案概览"

        headers = ['学号', '姓名', '批次', '班级', '政治身份', '材料齐全度', '转接状态']
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        completeness_map = {
            'complete': '齐全',
            'incomplete': '不齐全',
            'partial': '部分齐全',
            'pending': '待审核'
        }
        for user in users:
            ws.append([
                user['student_id'],
                user['name'],
                user['batch'] or '未分配',
                user['class_name'] or '未知',
                user['identity_names'] or '无',
                completeness_map.get(user['completeness'], user['completeness']),
                user['transfer_status'] or '未开始'
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'archive_overview_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error exporting archive: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/cultivators', methods=['GET', 'POST'])
@admin_required
def admin_cultivators():
    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'add':
                name = request.form['name']
                role = request.form['role']
                add_cultivator(name, role)
                log_operation(session['user_id'], session.get('student_id', '管理员'), '添加培养人', f'{name}({role})')
                flash('培养人添加成功', 'success')
            elif action == 'edit':
                cultivator_id = request.form['cultivator_id']
                name = request.form['name']
                role = request.form['role']
                update_cultivator(cultivator_id, name, role)
                log_operation(session['user_id'], session.get('student_id', '管理员'), '修改培养人', f'{name}({role})')
                flash('培养人更新成功', 'success')
            elif action == 'delete':
                cultivator_id = request.form['cultivator_id']
                delete_cultivator(cultivator_id)
                log_operation(session['user_id'], session.get('student_id', '管理员'), '删除培养人', f'ID:{cultivator_id}')
                flash('培养人删除成功', 'success')
        except Exception as e:
            flash(f'操作失败：{str(e)}', 'error')
            logger.error(f"Error in cultivators: {str(e)}")
        return redirect(url_for('admin_cultivators'))
    
    cultivators = get_all_cultivators()
    return render_template('admin_cultivators.html', cultivators=cultivators)

@app.route('/admin/update_materials', methods=['POST'])
@admin_required
def update_materials():
    try:
        user_id = request.form['user_id']
        material_id = request.form['material_id']
        status = request.form['status']
        update_material_status(user_id, material_id, status)
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error updating materials: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/update_announcement', methods=['POST'])
@admin_required
def update_announcement_route():
    try:
        content = request.form['content']
        update_date = request.form['update_date']
        update_announcement(content, update_date)
        log_operation(session['user_id'], session.get('student_id', '管理员'), '更新公告')
        flash('公告更新成功', 'success')
    except Exception as e:
        flash(f'公告更新失败：{str(e)}', 'error')
        logger.error(f"Error updating announcement: {str(e)}")
    return redirect(url_for('admin'))

@app.route('/admin/update_contact', methods=['POST'])
@admin_required
def update_contact():
    try:
        phone = request.form['phone']
        update_contact_info(phone)
        flash('联系电话更新成功', 'success')
    except Exception as e:
        flash(f'联系电话更新失败：{str(e)}', 'error')
        logger.error(f"Error updating contact: {str(e)}")
    return redirect(url_for('admin'))

@app.route('/admin/update_help', methods=['POST'])
@admin_required
def update_help():
    try:
        content = request.form['content']
        update_help_content(content)
        log_operation(session['user_id'], session.get('student_id', '管理员'), '更新帮助内容')
        flash('帮助内容更新成功', 'success')
    except Exception as e:
        flash(f'帮助内容更新失败：{str(e)}', 'error')
        logger.error(f"Error updating help content: {str(e)}")
    return redirect(url_for('admin'))

@app.route('/admin/operation_logs')
@admin_required
def operation_logs():
    result = search_operation_logs(page=1, per_page=20)
    return render_template('admin_operation_logs.html', 
                         logs=result['logs'],
                         total=result['total'],
                         page=result['page'],
                         total_pages=result['total_pages'],
                         actions=result['actions'])

@app.route('/admin/api/operation_logs')
@admin_required
def api_operation_logs():
    try:
        search = request.args.get('search', '').strip() or None
        action = request.args.get('action', '').strip() or None
        date_from = request.args.get('date_from', '').strip() or None
        date_to = request.args.get('date_to', '').strip() or None
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        result = search_operation_logs(
            search=search, action=action,
            date_from=date_from, date_to=date_to,
            page=page, per_page=per_page
        )
        return jsonify({'success': True, **result})
    except Exception as e:
        logger.error(f"Error searching operation logs: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/statistics')
@admin_required
def statistics():
    try:
        conn = get_db()
        with conn:
            # 身份分布
            identity_stats = conn.execute('''
                SELECT i.name, COUNT(DISTINCT ui.user_id) as count
                FROM identities i
                LEFT JOIN user_identities ui ON i.id = ui.identity_id
                LEFT JOIN users u ON ui.user_id = u.id AND u.is_admin = 0
                GROUP BY i.id, i.name
                ORDER BY i.id
            ''').fetchall()

            # 转接状态分布
            transfer_stats = conn.execute('''
                SELECT ts.transfer_status, COUNT(*) as count
                FROM transfer_status ts
                JOIN users u ON ts.user_id = u.id
                WHERE u.is_admin = 0
                GROUP BY ts.transfer_status
            ''').fetchall()

            # 材料齐全度分布
            completeness_stats = conn.execute('''
                SELECT
                    CASE
                        WHEN COUNT(um.status) = 0 THEN '未录入'
                        WHEN SUM(CASE WHEN um.status = '齐全' THEN 1 ELSE 0 END) = COUNT(um.status) THEN '齐全'
                        WHEN SUM(CASE WHEN um.status = '齐全' THEN 1 ELSE 0 END) = 0 THEN '不齐全'
                        ELSE '部分齐全'
                    END AS completeness,
                    COUNT(DISTINCT um.user_id) as count
                FROM users u
                LEFT JOIN user_materials um ON u.id = um.user_id
                WHERE u.is_admin = 0
                GROUP BY u.id
            ''').fetchall()

            # 汇总齐全度
            completeness_summary = {}
            for row in completeness_stats:
                key = row['completeness']
                completeness_summary[key] = completeness_summary.get(key, 0) + row['count']

            # 批次分布
            batch_stats = conn.execute('''
                SELECT COALESCE(batch, '未分配') as batch, COUNT(*) as count
                FROM users
                WHERE is_admin = 0
                GROUP BY batch
                ORDER BY count DESC
            ''').fetchall()

            # 班级分布
            class_stats = conn.execute('''
                SELECT COALESCE(class_name, '未知') as class_name, COUNT(*) as count
                FROM users
                WHERE is_admin = 0
                GROUP BY class_name
                ORDER BY count DESC
            ''').fetchall()

            # 总体统计
            total_users = conn.execute('SELECT COUNT(*) FROM users WHERE is_admin = 0').fetchone()[0]
            total_completed = conn.execute('''
                SELECT COUNT(*) FROM transfer_status ts
                JOIN users u ON ts.user_id = u.id
                WHERE u.is_admin = 0 AND ts.transfer_status = '已完成'
            ''').fetchone()[0]
            total_processing = conn.execute('''
                SELECT COUNT(*) FROM transfer_status ts
                JOIN users u ON ts.user_id = u.id
                WHERE u.is_admin = 0 AND ts.transfer_status = '处理中'
            ''').fetchone()[0]
            total_abandoned = conn.execute('''
                SELECT COUNT(*) FROM transfer_status ts
                JOIN users u ON ts.user_id = u.id
                WHERE u.is_admin = 0 AND ts.transfer_status = '已放弃'
            ''').fetchone()[0]
            avg_completeness = conn.execute('''
                SELECT AVG(
                    CASE WHEN total = 0 THEN 0
                    ELSE complete * 100.0 / total END
                ) FROM (
                    SELECT u.id,
                        SUM(CASE WHEN um.status = '齐全' THEN 1 ELSE 0 END) as complete,
                        COUNT(um.status) as total
                    FROM users u
                    LEFT JOIN user_materials um ON u.id = um.user_id
                    WHERE u.is_admin = 0
                    GROUP BY u.id
                )
            ''').fetchone()[0] or 0

        conn.close()
        return render_template('admin_statistics.html',
            identity_stats=[dict(r) for r in identity_stats],
            transfer_stats=[dict(r) for r in transfer_stats],
            completeness_summary=completeness_summary,
            batch_stats=[dict(r) for r in batch_stats],
            class_stats=[dict(r) for r in class_stats],
            total_users=total_users,
            total_completed=total_completed,
            total_processing=total_processing,
            total_abandoned=total_abandoned,
            avg_completeness=round(avg_completeness, 1)
        )
    except Exception as e:
        logger.error(f"Error loading statistics: {str(e)}")
        flash(f'加载统计数据失败：{str(e)}', 'error')
        return redirect(url_for('admin'))

# ==================== 党费缴纳记录管理 ====================

@app.route('/admin/fee_records')
@admin_required
def admin_fee_records():
    year = request.args.get('year', '').strip() or None
    month = request.args.get('month', '').strip() or None
    status = request.args.get('status', '').strip() or None
    page = int(request.args.get('page', 1))

    result = get_fee_records(year=year, month=month, status=status, page=page, per_page=20)
    statistics = get_fee_statistics(year=year)
    users = get_all_users()

    return render_template('admin_fee_records.html',
                         records=result['records'],
                         total=result['total'],
                         page=result['page'],
                         total_pages=result['total_pages'],
                         statistics=statistics,
                         users=[dict(u) for u in users if not u['is_admin']],
                         filter_year=year,
                         filter_month=month,
                         filter_status=status)

@app.route('/admin/fee_records/add', methods=['POST'])
@admin_required
def add_fee_record_route():
    try:
        user_id = int(request.form['user_id'])
        year = int(request.form['year'])
        month = int(request.form['month'])
        amount = float(request.form['amount'])
        status = request.form.get('status', '未缴纳')
        paid_date = request.form.get('paid_date') or None
        remark = request.form.get('remark', '').strip() or None

        add_fee_record(user_id, year, month, amount, status, paid_date, remark)
        log_operation(session['user_id'], session.get('student_id', '管理员'), '添加党费记录',
                     f'用户ID:{user_id}', f'{year}年{month}月 {amount}元')
        flash('党费记录添加成功', 'success')
    except sqlite3.IntegrityError:
        flash('该用户在此月份已有记录', 'error')
    except Exception as e:
        flash(f'添加失败：{str(e)}', 'error')
        logger.error(f"Error adding fee record: {str(e)}")
    return redirect(url_for('admin_fee_records'))

@app.route('/admin/fee_records/update', methods=['POST'])
@admin_required
def update_fee_record_route():
    try:
        record_id = int(request.form['record_id'])
        amount = float(request.form['amount']) if request.form.get('amount') else None
        status = request.form.get('status')
        paid_date = request.form.get('paid_date') or None
        remark = request.form.get('remark', '').strip() or None

        update_fee_record(record_id, amount=amount, status=status, paid_date=paid_date, remark=remark)
        log_operation(session['user_id'], session.get('student_id', '管理员'), '更新党费记录',
                     f'记录ID:{record_id}', f'状态:{status}')
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error updating fee record: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/fee_records/delete', methods=['POST'])
@admin_required
def delete_fee_record_route():
    try:
        record_id = int(request.form['record_id'])
        delete_fee_record(record_id)
        log_operation(session['user_id'], session.get('student_id', '管理员'), '删除党费记录', f'记录ID:{record_id}')
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting fee record: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/fee_records/batch_generate', methods=['POST'])
@admin_required
def batch_generate_fees():
    try:
        year = int(request.form['year'])
        months = [int(m) for m in request.form.getlist('months')]
        amount = float(request.form['amount'])

        if not months:
            flash('请至少选择一个月份', 'error')
            return redirect(url_for('admin_fee_records'))

        count = batch_generate_fee_records(year, months, amount)
        log_operation(session['user_id'], session.get('student_id', '管理员'), '批量生成党费记录',
                     f'{year}年 {len(months)}个月', f'生成{count}条')
        flash(f'成功生成 {count} 条党费记录', 'success')
    except Exception as e:
        flash(f'批量生成失败：{str(e)}', 'error')
        logger.error(f"Error batch generating fees: {str(e)}")
    return redirect(url_for('admin_fee_records'))

@app.route('/admin/fee_records/export')
@admin_required
def export_fee_records():
    try:
        year = request.args.get('year', '').strip() or None
        month = request.args.get('month', '').strip() or None
        status = request.args.get('status', '').strip() or None

        result = get_fee_records(year=year, month=month, status=status, page=1, per_page=10000)

        wb = Workbook()
        ws = wb.active
        ws.title = "党费缴纳记录"

        headers = ['学号', '姓名', '班级', '批次', '年份', '月份', '金额(元)', '缴纳状态', '缴纳日期', '备注']
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for record in result['records']:
            ws.append([
                record['student_id'],
                record['name'],
                record['class_name'],
                record['batch'] or '未分配',
                record['year'],
                record['month'],
                record['amount'],
                record['status'],
                record['paid_date'] or '',
                record['remark'] or ''
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = (max_length + 2) * 1.2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'fee_records_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error exporting fee records: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/api/fee_records')
@admin_required
def api_fee_records():
    try:
        search = request.args.get('search', '').strip()
        year = request.args.get('year', '').strip() or None
        month = request.args.get('month', '').strip() or None
        status = request.args.get('status', '').strip() or None
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        result = get_fee_records(year=year, month=month, status=status, page=page, per_page=per_page)

        if search:
            search_lower = search.lower()
            result['records'] = [r for r in result['records']
                if search_lower in r['student_id'].lower()
                or search_lower in r['name'].lower()
                or search_lower in (r['class_name'] or '').lower()]
            result['total'] = len(result['records'])

        return jsonify({'success': True, **result})
    except Exception as e:
        logger.error(f"Error fetching fee records: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=8888)